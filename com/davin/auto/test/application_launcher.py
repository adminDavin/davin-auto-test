import logging
import json
from openpyxl import load_workbook
import requests


logger = logging.getLogger()
logger.setLevel("DEBUG")
BASIC_FORMAT = "%(asctime)s:%(levelname)s:%(message)s"
DATE_FORMAT = '%Y-%m-%d %H:%M:%S'
formatter = logging.Formatter(BASIC_FORMAT, DATE_FORMAT)
chlr = logging.StreamHandler() # 输出到控制台的handler
chlr.setFormatter(formatter)
logger.addHandler(chlr)


def read_file_data(file_name):
    with open(file_name, "r") as f:
        return f.read()


opjson = json.loads(read_file_data('resources/opjson.json'))


class AutoTest:
    def __init__(self, label, url, is_exc, method, has_header, headers, data_flag, relate_label, return_field, relate_body_field):
        self.label = label
        self.url = url
        self.is_exc = is_exc
        self.method = method
        self.has_header = has_header
        print(headers)
        self.headers = json.loads(headers)
        self.data_flag = data_flag
        self.relate_label = relate_label
        self.return_field = return_field
        self.relate_body_field = relate_body_field

    def has_related(self):
        if self.relate_label is None:
            return None
        else:
            return self.relate_label

    def handle_related(self, relate, tests):
        if relate is not None:
            relate_test = tests[relate]
            data = relate_test.execute_test(tests)
            if 'login' == relate_test.data_flag:
                authorization = data.headers['Authorization']
                self.headers['authorization'] = authorization
            else:
                self.headers['authorization'] = relate_test.headers['authorization']
            return data
        return None

    def request_get(self, tests):
        relate = self.has_related()
        data = self.handle_related(relate, tests)
        if data is not None:
            pass

        request_body = opjson[self.data_flag]
        if self.relate_body_field is not None:
            request_body[self.relate_body_field] = data
        r = requests.get(self.url, headers=self.headers, data=request_body)
        return r

    def request_post(self, tests):
        relate = self.has_related()
        data = self.handle_related(relate, tests)
        if data is not None:
            pass
        request_body = opjson[self.data_flag]
        if self.relate_body_field is not None:
            request_body[self.relate_body_field] = data
        r = requests.post(self.url, headers=self.headers, data=request_body)
        print(r.headers, json.loads(r.content))
        return r

    def execute_test(self, tests):
        if 'yes' == self.is_exc:
            if 'post' == self.method:
                return self.request_post(tests)
            else:
                return self.request_get(tests)


def load_tests(file, sheet):
    wb = load_workbook(file)
    wb.guess_types = True
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows():
        rows.append(row)

    return rows


def prepare_tests(rows):
    tests = {}
    for i, row in enumerate(rows):
        # print(row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12])
        if i == 0:
            continue
        label, url, is_exc, method = row[0].value, row[2].value, row[3].value, row[4].value
        has_header, headers, data_flag, relate_label = row[5].value, row[6].value, row[7].value, row[8].value
        return_field, relate_body_field = row[9].value, row[10].value
        if url is None:
            continue

        auto_test = AutoTest(label, url, is_exc, method, has_header, headers, data_flag, relate_label, return_field, relate_body_field)
        tests[i] = auto_test
    return tests


def main():
    rows = load_tests('resources/a.xlsx', 'Sheet1')
    tests = prepare_tests(rows)
    print(tests)
    for i in tests:
        tests[i].execute_test(tests)
    logger.info('begin start autotest')


if __name__ == '__main__':
    main()